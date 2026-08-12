from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


LEVELS = ("A1", "A2", "A2+", "B1", "B1+", "B2")
DEFAULT_HINT_1 = (
    "Osserva l’ampiezza del compito, il tipo di contenuto e le condizioni "
    "indicate nel descrittore."
)
DEFAULT_HINT_2 = (
    "Confronta il descrittore con i livelli vicini disponibili per questa "
    "scala, senza basarti su una sola parola."
)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _feedback_by_text(sample_path: Path) -> dict[str, dict[str, Any]]:
    if not sample_path.exists():
        return {}
    payload = json.loads(sample_path.read_text(encoding="utf-8"))
    return {
        _text(item.get("descriptor_text")): item
        for item in payload
        if _text(item.get("descriptor_text"))
    }


def _navigation_fields(
    schema: str,
    modality: str,
    activity: str,
    scale: str,
) -> tuple[str, str, str]:
    navigation_schema = schema or "Catalogo CEFR"
    if "lingua dei segni" in schema.casefold() or "lingue dei segni" in schema.casefold():
        competence_names = {
            "competenza linguistica": "Linguistica",
            "competenza sociolinguistica": "Sociolinguistica",
            "competenza pragmatica": "Pragmatica",
        }
        navigation_modality = competence_names.get(
            activity.casefold(), activity or "Altre competenze"
        )
        navigation_activity = modality or "Scale generali"
        return navigation_schema, navigation_modality, navigation_activity
    return (
        navigation_schema,
        modality or activity or "Altre scale",
        activity or "Scale disponibili",
    )


def build_catalog(
    source: Path,
    sample_path: Path,
    *,
    expected_rows: int = 831,
) -> tuple[list[dict], dict]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    feedback = _feedback_by_text(sample_path)
    rows: list[dict[str, Any]] = []
    ignored: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    source_rows = 0
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, max_col=7, values_only=True), start=2
    ):
        source_rows += 1
        source_id, schema, modality, activity, scale, level, descriptor = values
        source_id = _text(source_id)
        source_schema = _text(schema)
        source_modality = _text(modality)
        source_activity = _text(activity)
        scale = _text(scale)
        level = _text(level).upper()
        descriptor = _text(descriptor)
        if not scale or level not in LEVELS or not descriptor:
            ignored.append({"row": row_number, "reason": "E, F o G non validi"})
            continue
        if descriptor.casefold() == "nessun descrittore":
            raise ValueError(f"Riga {row_number}: reintrodotto 'Nessun descrittore'.")
        descriptor_id = f"SRC-{source_id}"
        if not source_id or descriptor_id in seen_ids:
            raise ValueError(
                f"Riga {row_number}: identificatore sorgente mancante o duplicato."
            )
        seen_ids.add(descriptor_id)
        editorial = feedback.get(descriptor, {})
        # B-C-D organizzano la navigazione, ma nella fonte C o D possono
        # legittimamente essere vuote. Le etichette di navigazione colmano
        # soltanto quei vuoti; i valori originali restano nei campi source_*.
        (
            navigation_schema,
            navigation_modality,
            navigation_activity,
        ) = _navigation_fields(
            source_schema,
            source_modality,
            source_activity,
            scale,
        )
        rows.append(
            {
                "descriptor_id": descriptor_id,
                "source_row_id": source_id,
                "schema": navigation_schema,
                "modality": navigation_modality,
                "activity": navigation_activity,
                "source_schema": source_schema,
                "source_modality": source_modality,
                "source_activity": source_activity,
                "scale": scale,
                "correct_level": level,
                "descriptor_text": descriptor,
                "rationale": editorial.get("rationale")
                or (
                    f"Il livello corretto è {level}, come indicato nel "
                    "catalogo sorgente revisionato."
                ),
                "hint_1": editorial.get("hint_1") or DEFAULT_HINT_1,
                "hint_2": editorial.get("hint_2") or DEFAULT_HINT_2,
                "language": "it",
                "source": source.name,
                "source_version": "2026-07-30",
                "license_or_permission": (
                    "Da verificare prima della pubblicazione pubblica"
                ),
                "content_version": "catalogo-completo-2026-07-30",
                "status": "approved",
                "active": True,
            }
        )

    report = {
        "source": str(source),
        "worksheet": sheet.title,
        "source_rows": source_rows,
        "catalog_rows": len(rows),
        "ignored_rows": ignored,
        "level_counts": dict(Counter(item["correct_level"] for item in rows)),
        "blank_source_schema": sum(not item["source_schema"] for item in rows),
        "blank_source_modality": sum(
            not item["source_modality"] for item in rows
        ),
        "blank_source_activity": sum(
            not item["source_activity"] for item in rows
        ),
        "unique_scales": len({item["scale"] for item in rows}),
    }
    if len(rows) != expected_rows or ignored:
        raise ValueError(
            "Il catalogo pulito deve produrre esattamente "
            f"{expected_rows} righe valide. "
            f"Ottenute {len(rows)}; ignorate {len(ignored)}."
        )
    return rows, report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Converte l’Excel CEFR pulito nel catalogo JSON privato."
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--sample",
        type=Path,
        default=Path(__file__).parents[1] / "space/data/catalog.sample.json",
    )
    args = parser.parse_args()
    catalog, report = build_catalog(args.source, args.sample)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    report_path = args.output.with_suffix(".report.json")
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Creati {len(catalog)} esercizi in {args.output}")
    print(f"Rapporto: {report_path}")


if __name__ == "__main__":
    main()
